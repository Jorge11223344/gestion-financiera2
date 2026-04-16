"""
Importador de cartolas bancarias.
Soporta:
  - Global66                  → CSV separador punto y coma (;), columna "Tipo" nativa
  - Banco de Chile / Edwards  → Excel (.xlsx / .xls)
  - CSV genérico              → fallback para Santander, BCI y otros

El banco se detecta automáticamente por el contenido y nombre del archivo.
Global66 tiene parser especializado que usa la columna "Tipo" del CSV para
clasificar movimientos con mayor precisión.
"""

import csv
import io
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass, field
from typing import List, Optional, Tuple


# ─────────────────────────────────────────────
# Estructuras de datos
# ─────────────────────────────────────────────

@dataclass
class FilaCartola:
    fecha: date
    descripcion: str
    cargo: Decimal
    abono: Decimal
    saldo: Optional[Decimal]
    documento: str = ""
    tipo_banco: str = ""    # columna "Tipo" nativa del banco (útil en Global66)
    fila_origen: int = 0
    moneda: str = "CLP"
    tipo_cambio: Optional[Decimal] = None
    tercero: str = ""
    pais: str = ""


@dataclass
class ResultadoImportacion:
    filas: List[FilaCartola] = field(default_factory=list)
    errores: List[str] = field(default_factory=list)
    advertencias: List[str] = field(default_factory=list)
    banco_detectado: str = "Desconocido"
    total_filas_archivo: int = 0
    total_filas_validas: int = 0


# ─────────────────────────────────────────────
# Utilidades de parseo
# ─────────────────────────────────────────────

def parse_fecha(valor) -> Optional[date]:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    s = str(valor).strip().split(" ")[0].split("T")[0]
    for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
                "%Y-%m-%d", "%Y/%m/%d", "%d %b %Y"]:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_monto(valor) -> Decimal:
    if valor is None or str(valor).strip() in ("", "-", "--", "N/A"):
        return Decimal("0")
    if isinstance(valor, (int, float)):
        return Decimal(str(abs(valor)))
    s = str(valor).strip().replace("$", "").replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        partes = s.split(",")
        s = s.replace(",", ".") if len(partes) == 2 and len(partes[1]) <= 2 else s.replace(",", "")
    elif "." in s:
        partes = s.split(".")
        if len(partes) > 2 or (len(partes) == 2 and len(partes[1]) == 3):
            s = s.replace(".", "")
    s = s.replace("(", "-").replace(")", "")
    try:
        return abs(Decimal(s))
    except InvalidOperation:
        return Decimal("0")


def decodificar(archivo_bytes: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return archivo_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("No se pudo decodificar el archivo")


def es_fila_encabezado(valores: list) -> bool:
    texto = " ".join(str(v).lower() for v in valores if v)
    kw = ["fecha", "descripcion", "descripción", "cargo", "abono", "saldo",
          "tipo", "detalle", "referencia", "monto", "débito", "crédito"]
    return sum(1 for k in kw if k in texto) >= 2


def detectar_columnas(fila: list) -> dict:
    mapa = {}
    for i, celda in enumerate(fila):
        t = str(celda).lower().strip()
        if any(p in t for p in ["fecha", "date"]):
            mapa.setdefault("fecha", i)
        elif t in ["tipo", "tipo transaccion", "tipo transacción", "tipo movimiento"]:
            mapa.setdefault("tipo_banco", i)
        elif any(p in t for p in ["descripcion", "descripción", "detalle", "glosa", "concepto"]):
            mapa.setdefault("descripcion", i)
        elif any(p in t for p in ["cargo", "débito", "debito", "retiro", "egreso"]):
            mapa.setdefault("cargo", i)
        elif any(p in t for p in ["abono", "crédito", "credito", "depósito", "deposito", "ingreso"]):
            mapa.setdefault("abono", i)
        elif any(p in t for p in ["saldo", "balance"]):
            mapa.setdefault("saldo", i)
        elif any(p in t for p in ["referencia", "documento", "n°", "folio", "ref"]):
            mapa.setdefault("documento", i)
    return mapa


def detectar_banco(texto: str, nombre: str = "") -> str:
    t, n = texto.lower(), nombre.lower()
    if "global66" in t or "global66" in n or "global 66" in t:
        return "Global66"
    # Archivos Global66 tienen nombre "extracto_movimientos_start_..."
    if "extracto_movimientos" in n:
        return "Global66"
    if "banco de chile" in t or "edwards" in t or "bancochile" in n:
        return "Banco de Chile / Edwards"
    if "santander" in t:
        return "Santander"
    if "bci" in t or "bci" in n:
        return "BCI"
    if "scotiabank" in t:
        return "Scotiabank"
    if "itau" in t or "itaú" in t:
        return "Itaú"
    return "CSV Genérico"


# ─────────────────────────────────────────────
# Clasificador Global66 (usa columna Tipo nativa)
# ─────────────────────────────────────────────

MAPA_GLOBAL66 = {
    "transferencia recibida": "ingreso_banco",
    "abono":                  "ingreso_banco",
    "deposito":               "ingreso_banco",
    "depósito":               "ingreso_banco",
    "conversion de divisas":  "conversion_divisa",
    "conversión de divisas":  "conversion_divisa",
    "transferencia enviada":  "egreso_banco",
    "cargo":                  "egreso_banco",
    "retiro":                 "egreso_banco",
    "comision":               "egreso_banco",
    "comisión":               "egreso_banco",
    "mantencion":             "egreso_banco",
    "mantención":             "egreso_banco",
    "saldo inicial":          None,   # ignorar
    "saldo anterior":         None,
}


def clasificar_global66(tipo_banco: str, descripcion: str, es_cargo: bool) -> Optional[str]:
    tb = tipo_banco.lower().strip()
    desc = descripcion.lower()

    # Para egresos: primero revisar descripción en busca de categorías específicas,
    # ya que "Transferencia enviada" puede ser remuneración, impuesto, etc.
    if es_cargo:
        if any(p in desc for p in ["remuneracion", "remuneración", "sueldo", "salario", "planilla"]):
            return "remuneracion"
        if any(p in desc for p in ["sii", "impuesto", "iva mensual", "renta", "patente", "tesoreria"]):
            return "impuesto"
        if any(p in desc for p in ["cuota prestamo", "cuota crédito", "credito hipotecario", "leasing"]):
            return "pago_prestamo"

    # Luego usar el tipo nativo del banco
    if tb in MAPA_GLOBAL66:
        return MAPA_GLOBAL66[tb]
    for clave, tipo in MAPA_GLOBAL66.items():
        if clave in tb:
            return tipo

    return "egreso_banco" if es_cargo else "ingreso_banco"


# ─────────────────────────────────────────────
# Parser Global66 XLSX (formato nuevo)
# ─────────────────────────────────────────────
# Global66 exporta archivos con extensión .xls que son XLSX reales (ZIP).
# Columnas: Tipo de transacción | Fecha | Monto debitado | Monto acreditado
#           | Costo tipo cambio | ID Fees | Últimos 4 dígitos | Nombre tercero
#           | DNI | N° cuenta | País | Tipo de cambio | ID transacción | Comentario

MAPA_TIPO_GLOBAL66_XLSX = {
    "conversión de divisas":        "conversion_divisa",
    "conversion de divisas":        "conversion_divisa",
    "intereses abonados":           "otro_ingreso",
    "envío a cuenta bancaria":      "egreso_banco",
    "envio a cuenta bancaria":      "egreso_banco",
    "comisión envío":               "egreso_banco",
    "comision envio":               "egreso_banco",
    "compra en":                    "egreso_banco",    # prefijo
}


def _clasificar_tipo_global66_xlsx(tipo: str, tiene_debito: bool, nombre_tercero: str = "", comentario: str = "") -> str:
    t = (tipo or "").lower().strip()
    tercero = (nombre_tercero or "").lower().strip()
    comentario = (comentario or "").lower().strip()

    if t.startswith("recibido de"):
        if "jimacomex" in tercero or "jimacomex" in comentario:
            return "transferencia_interna"
        return "ingreso_banco"

    if "conversión de divisas" in t or "conversion de divisas" in t:
        return "conversion_divisa"

    if t.startswith("envío a") or t.startswith("envio a"):
        return "egreso_banco"

    if t.startswith("compra en"):
        return "egreso_banco"

    for clave, tipo_mov in MAPA_TIPO_GLOBAL66_XLSX.items():
        if t == clave or t.startswith(clave):
            return tipo_mov

    return "egreso_banco" if tiene_debito else "ingreso_banco"


def _parsear_global66_xlsx(archivo_bytes: bytes) -> ResultadoImportacion:
    """Parser para el XLSX real que exporta Global66 (extensión .xls)."""
    import zipfile
    import html as html_mod
    import re as re_mod

    resultado = ResultadoImportacion()
    resultado.banco_detectado = "Global66"

    try:
        zf = zipfile.ZipFile(io.BytesIO(archivo_bytes))
        sheet_xml = zf.read('xl/worksheets/sheet1.xml').decode('utf-8')
    except Exception as e:
        resultado.errores.append(f"No se pudo abrir el archivo Global66: {e}")
        return resultado

    # Extraer filas y celdas inline del XML
    rows = re_mod.findall(r'<row[^>]*>(.*?)</row>', sheet_xml, re_mod.DOTALL)
    resultado.total_filas_archivo = len(rows)

    # Detectar moneda de la cuenta (fila 0: "Movimientos de cuenta USD/CLP")
    moneda = "CLP"
    if rows:
        primera = re_mod.search(r'<is><t[^>]*>(.*?)</t>', rows[0])
        if primera and "USD" in primera.group(1).upper():
            moneda = "USD"

    resultado.advertencias.append(f"Cuenta Global66 en {moneda}")

    def extraer_celdas(row_xml):
        cells = re_mod.findall(r'<c[^>]*>(.*?)</c>', row_xml, re_mod.DOTALL)
        valores = []
        for cell in cells:
            v = re_mod.search(r'<is><t[^>]*>(.*?)</t>', cell)
            if not v:
                v = re_mod.search(r'<v>(.*?)</v>', cell)
            valores.append(html_mod.unescape(v.group(1)) if v else "")
        return valores

    # Fila 2 es el encabezado (índice 2)
    # Columnas fijas del formato Global66 XLSX:
    # 0: Tipo de transacción
    # 1: Fecha de la transacción
    # 2: Monto debitado
    # 3: Monto acreditado
    # 4: Costo de tipo de cambio
    # 5: ID Fees Asociados
    # 6: Últimos 4 dígitos tarjeta
    # 7: Nombre tercero o Comercio
    # 8: DNI del tercero
    # 9: Número de cuenta del tercero
    # 10: País destino/tercero
    # 11: Tipo de cambio
    # 12: ID de la transacción
    # 13: Comentario

    enc_idx = None
    for i, row in enumerate(rows):
        vals = extraer_celdas(row)
        if vals and "tipo" in vals[0].lower() and "fecha" in " ".join(vals).lower():
            enc_idx = i
            break

    if enc_idx is None:
        resultado.errores.append(
            "No se encontró el encabezado en el archivo Global66. "
            "Descarga desde: Cuenta → Movimientos → Exportar."
        )
        return resultado

    for i, row in enumerate(rows[enc_idx + 1:], enc_idx + 2):
        vals = extraer_celdas(row)
        if not any(vals):
            continue

        # Columna fecha (índice 1)
        fecha_str = vals[1].strip() if len(vals) > 1 else ""
        fecha = parse_fecha(fecha_str.split(" ")[0] if fecha_str else None)
        if fecha is None:
            continue

        tipo_transaccion = vals[0].strip() if len(vals) > 0 else ""
        monto_debito_str  = vals[2].strip() if len(vals) > 2 else ""
        monto_credito_str = vals[3].strip() if len(vals) > 3 else ""
        nombre_tercero    = vals[7].strip() if len(vals) > 7 else ""
        comentario        = vals[13].strip() if len(vals) > 13 else ""
        id_transaccion    = vals[12].strip() if len(vals) > 12 else ""
        pais              = vals[10].strip() if len(vals) > 10 else ""
        tc_str            = vals[11].strip() if len(vals) > 11 else ""
        tipo_cambio       = parse_monto(tc_str) if tc_str else None

        cargo = parse_monto(monto_debito_str)
        abono = parse_monto(monto_credito_str)

        if cargo == 0 and abono == 0:
            continue

        tiene_debito = cargo > 0
        tipo_mov = _clasificar_tipo_global66_xlsx(tipo_transaccion, tiene_debito, nombre_tercero, comentario)

        # Construir descripción legible
        if nombre_tercero:
            descripcion = f"{tipo_transaccion}: {nombre_tercero}"
        else:
            descripcion = tipo_transaccion
        if comentario:
            descripcion += f" | {comentario}"
        descripcion = descripcion[:250]

        resultado.filas.append(FilaCartola(
            fecha=fecha,
            descripcion=descripcion,
            cargo=cargo,
            abono=abono,
            saldo=None,
            documento=id_transaccion,
            tipo_banco=tipo_transaccion,
            fila_origen=i,
            moneda=moneda,
            tipo_cambio=tipo_cambio if tipo_cambio and tipo_cambio > 0 else None,
            tercero=nombre_tercero,
            pais=pais,
        ))

    resultado.total_filas_validas = len(resultado.filas)
    return resultado


# ─────────────────────────────────────────────
# Parser Global66 (especializado - enrutador)
# ─────────────────────────────────────────────

def parsear_global66(archivo_bytes: bytes, nombre_archivo: str = "") -> ResultadoImportacion:
    # Global66 puede exportar en dos formatos:
    # 1. XLSX real con extensión .xls (magic PK = ZIP) — formato nuevo
    # 2. CSV con separador punto y coma — formato antiguo
    magic_zip = b'PK\x03\x04'
    if archivo_bytes[:4] == magic_zip:
        return _parsear_global66_xlsx(archivo_bytes)

    # --- CSV legacy ---
    resultado = ResultadoImportacion()
    resultado.banco_detectado = "Global66"

    try:
        texto = decodificar(archivo_bytes)
    except ValueError as e:
        resultado.errores.append(str(e))
        return resultado

    reader = csv.reader(io.StringIO(texto), delimiter=";")
    todas_filas = list(reader)
    resultado.total_filas_archivo = len(todas_filas)

    # Buscar encabezado
    mapa_columnas = {}
    fila_enc_idx = None
    for i, fila in enumerate(todas_filas):
        if es_fila_encabezado(fila):
            mapa = detectar_columnas(fila)
            if "fecha" in mapa:
                mapa_columnas = mapa
                fila_enc_idx = i
                break

    if not mapa_columnas:
        resultado.errores.append(
            "No se detectó encabezado en el CSV de Global66. "
            "Exporta desde: Cuenta → Movimientos → Descargar CSV."
        )
        return resultado

    for i, fila in enumerate(todas_filas[fila_enc_idx + 1:], fila_enc_idx + 2):
        if not any(f.strip() for f in fila):
            continue

        fecha = parse_fecha(fila[mapa_columnas["fecha"]] if len(fila) > mapa_columnas.get("fecha", 99) else None)
        if fecha is None:
            continue

        tipo_banco_idx = mapa_columnas.get("tipo_banco")
        tipo_banco = fila[tipo_banco_idx].strip() if tipo_banco_idx is not None and len(fila) > tipo_banco_idx else ""

        desc_idx = mapa_columnas.get("descripcion")
        doc_idx  = mapa_columnas.get("documento")
        desc_raw  = fila[desc_idx].strip() if desc_idx is not None and len(fila) > desc_idx else ""
        referencia = fila[doc_idx].strip() if doc_idx is not None and len(fila) > doc_idx else ""

        if not desc_raw:
            desc_raw = tipo_banco or "Movimiento Global66"

        descripcion = f"{desc_raw} [{referencia}]" if referencia and referencia not in desc_raw else desc_raw

        cargo_idx = mapa_columnas.get("cargo")
        abono_idx = mapa_columnas.get("abono")
        cargo = parse_monto(fila[cargo_idx]) if cargo_idx is not None and len(fila) > cargo_idx else Decimal("0")
        abono = parse_monto(fila[abono_idx]) if abono_idx is not None and len(fila) > abono_idx else Decimal("0")

        saldo_idx = mapa_columnas.get("saldo")
        saldo = parse_monto(fila[saldo_idx]) if saldo_idx is not None and len(fila) > saldo_idx else None

        if cargo == 0 and abono == 0:
            continue

        resultado.filas.append(FilaCartola(
            fecha=fecha,
            descripcion=descripcion[:250],
            cargo=cargo,
            abono=abono,
            saldo=saldo,
            documento=referencia,
            tipo_banco=tipo_banco,
            fila_origen=i,
        ))

    resultado.total_filas_validas = len(resultado.filas)
    return resultado


# ─────────────────────────────────────────────
# Parser Excel (Banco de Chile y otros)
# ─────────────────────────────────────────────

def _es_csv_disfrazado(archivo_bytes: bytes) -> bool:
    """
    El Banco de Chile exporta archivos .xls que en realidad son CSV con
    separador punto y coma. Se detectan porque sus primeros bytes son texto
    legible (no el magic 0xD0CF del formato binario OLE2 de Excel).
    """
    magic_xls  = b'\xd0\xcf\x11\xe0'   # Excel binario real (OLE2)
    magic_xlsx = b'PK\x03\x04'          # ZIP → xlsx
    inicio = archivo_bytes[:8]
    return not (inicio[:4] == magic_xls or inicio[:4] == magic_xlsx)


def _parsear_csv_banco_chile(archivo_bytes: bytes) -> ResultadoImportacion:
    """
    Parser para el CSV con ; que exporta Banco de Chile con extensión .xls.
    Formato de montos: +0000200000  (sin decimales, signo + o vacío para abono,
    signo + para cargo — la columna ya indica si es cargo o abono).
    """
    resultado = ResultadoImportacion()
    resultado.banco_detectado = "Banco de Chile / Edwards"

    try:
        texto = decodificar(archivo_bytes)
    except ValueError as e:
        resultado.errores.append(str(e))
        return resultado

    lineas = texto.splitlines()
    resultado.total_filas_archivo = len(lineas)

    # La primera línea suele ser el encabezado de la empresa, la segunda el
    # encabezado de columnas. Buscamos la fila que contenga "Fecha".
    fila_enc_idx = None
    for i, linea in enumerate(lineas):
        if "fecha" in linea.lower() and ";" in linea:
            fila_enc_idx = i
            break

    if fila_enc_idx is None:
        resultado.errores.append(
            "No se encontró encabezado en el archivo del Banco de Chile. "
            "Descarga la cartola desde: Empresas → Cuenta Corriente → Cartola → Exportar."
        )
        return resultado

    encabezado = [c.strip().lower() for c in lineas[fila_enc_idx].split(";")]

    # Mapear columnas flexiblemente
    def col(nombres):
        for n in nombres:
            for i, h in enumerate(encabezado):
                if n in h:
                    return i
        return None

    idx_fecha  = col(["fecha"])
    idx_detalle = col(["detalle", "descripcion", "glosa", "concepto"])
    idx_cargo  = col(["cheque", "cargo", "débito", "debito", "retiro"])
    idx_abono  = col(["deposito", "depósito", "abono", "crédito", "credito"])
    idx_saldo  = col(["saldo"])
    idx_doc    = col(["docto", "documento", "n°", "folio", "ref"])

    if idx_fecha is None:
        resultado.errores.append("No se encontró columna de fecha en el archivo.")
        return resultado

    def parse_monto_banco_chile(valor: str) -> Decimal:
        """Convierte '+0000200000' o '00000000000' a Decimal."""
        if not valor:
            return Decimal("0")
        s = str(valor).strip().replace("+", "").replace("-", "").replace(".", "").replace(",", "")
        try:
            return Decimal(s)
        except Exception:
            return Decimal("0")

    for i, linea in enumerate(lineas[fila_enc_idx + 1:], fila_enc_idx + 2):
        if not linea.strip() or ";" not in linea:
            continue
        cols = linea.split(";")

        fecha = parse_fecha(cols[idx_fecha].strip() if idx_fecha is not None and len(cols) > idx_fecha else None)
        if fecha is None:
            continue

        detalle = cols[idx_detalle].strip() if idx_detalle is not None and len(cols) > idx_detalle else "Movimiento bancario"
        if not detalle:
            detalle = "Movimiento bancario"

        cargo = parse_monto_banco_chile(cols[idx_cargo]) if idx_cargo is not None and len(cols) > idx_cargo else Decimal("0")
        abono = parse_monto_banco_chile(cols[idx_abono]) if idx_abono is not None and len(cols) > idx_abono else Decimal("0")
        saldo = parse_monto_banco_chile(cols[idx_saldo]) if idx_saldo is not None and len(cols) > idx_saldo else None
        documento = cols[idx_doc].strip() if idx_doc is not None and len(cols) > idx_doc else ""
        if documento in ("0", "00000000000", ""):
            documento = ""

        if cargo == 0 and abono == 0:
            continue

        resultado.filas.append(FilaCartola(
            fecha=fecha,
            descripcion=detalle[:250],
            cargo=cargo,
            abono=abono,
            saldo=saldo,
            documento=documento,
            fila_origen=i,
        ))

    resultado.total_filas_validas = len(resultado.filas)
    return resultado


def parsear_excel(archivo_bytes: bytes, nombre_archivo: str = "") -> ResultadoImportacion:
    resultado = ResultadoImportacion()

    # Caso 1: Global66 exporta XLSX real con extensión .xls
    nombre_lower = nombre_archivo.lower()
    magic_zip = b'PK\x03\x04'
    if archivo_bytes[:4] == magic_zip:
        if "extracto_movimientos" in nombre_lower or "global66" in nombre_lower:
            return _parsear_global66_xlsx(archivo_bytes)
        try:
            import zipfile
            zf = zipfile.ZipFile(io.BytesIO(archivo_bytes))
            sheet = zf.read('xl/worksheets/sheet1.xml').decode('utf-8')
            if "Movimientos de cuenta" in sheet or "Tipo de transacci" in sheet:
                return _parsear_global66_xlsx(archivo_bytes)
        except Exception:
            pass

    # Caso 2: Banco de Chile exporta CSV con punto y coma con extensión .xls
    if _es_csv_disfrazado(archivo_bytes):
        return _parsear_csv_banco_chile(archivo_bytes)

    # Detectar formato por extensión y magic bytes
    nombre_lower = nombre_archivo.lower()
    es_xls = nombre_lower.endswith(".xls") or archivo_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'

    todas_filas = []

    if es_xls:
        # Formato antiguo .xls binario → usar xlrd
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=archivo_bytes)
            ws = wb.sheet_by_index(0)
            for row_idx in range(ws.nrows):
                fila = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    # Convertir tipos de celda xlrd a valores Python
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                            fila.append(datetime(*dt_tuple[:6]).date() if dt_tuple[3:] == (0,0,0) else datetime(*dt_tuple[:6]))
                        except Exception:
                            fila.append(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        # Preservar como número (evita que 1234567.0 se convierta en string)
                        v = cell.value
                        fila.append(int(v) if v == int(v) else v)
                    elif cell.ctype == xlrd.XL_CELL_EMPTY:
                        fila.append(None)
                    else:
                        fila.append(cell.value)
                todas_filas.append(tuple(fila))
        except ImportError:
            resultado.errores.append(
                "Archivo .xls detectado pero falta la librería xlrd. "
                "Ejecuta: pip install xlrd==2.0.1"
            )
            return resultado
        except Exception as e:
            resultado.errores.append(f"No se pudo abrir el archivo .xls: {e}")
            return resultado
    else:
        # Formato moderno .xlsx → usar openpyxl
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes), data_only=True)
            ws = wb.active
            todas_filas = list(ws.iter_rows(values_only=True))
        except Exception as e:
            resultado.errores.append(f"No se pudo abrir el Excel: {e}")
            return resultado

    resultado.total_filas_archivo = len(todas_filas)

    texto_cab = " ".join(str(c).lower() for fila in todas_filas[:6] for c in fila if c)
    resultado.banco_detectado = detectar_banco(texto_cab, nombre_archivo)

    fila_enc_idx = None
    mapa_columnas = {}
    for i, fila in enumerate(todas_filas):
        if len([v for v in fila if v is not None]) >= 3 and es_fila_encabezado(fila):
            mapa = detectar_columnas(fila)
            if "fecha" in mapa and ("cargo" in mapa or "abono" in mapa):
                fila_enc_idx = i
                mapa_columnas = mapa
                break

    if fila_enc_idx is None:
        resultado.advertencias.append("Encabezado no detectado. Usando detección automática.")
        mapa_columnas, fila_enc_idx = _detectar_columnas_heuristico(todas_filas)

    if not mapa_columnas or "fecha" not in mapa_columnas:
        resultado.errores.append(
            "No se identificaron columnas. Exporta la cartola en formato estándar desde el banco."
        )
        return resultado

    for i, fila in enumerate(todas_filas[fila_enc_idx + 1:], fila_enc_idx + 2):
        if not any(fila):
            continue
        valores = list(fila)
        fecha = parse_fecha(valores[mapa_columnas["fecha"]] if len(valores) > mapa_columnas.get("fecha", 99) else None)
        if fecha is None:
            continue

        desc_idx = mapa_columnas.get("descripcion")
        descripcion = str(valores[desc_idx]).strip() if desc_idx is not None and len(valores) > desc_idx else "Movimiento bancario"
        if not descripcion or descripcion.lower() in ("none", ""):
            descripcion = "Movimiento bancario"

        cargo_idx = mapa_columnas.get("cargo")
        abono_idx = mapa_columnas.get("abono")
        cargo = parse_monto(valores[cargo_idx]) if cargo_idx is not None and len(valores) > cargo_idx else Decimal("0")
        abono = parse_monto(valores[abono_idx]) if abono_idx is not None and len(valores) > abono_idx else Decimal("0")

        saldo_idx = mapa_columnas.get("saldo")
        saldo = parse_monto(valores[saldo_idx]) if saldo_idx is not None and len(valores) > saldo_idx else None

        doc_idx = mapa_columnas.get("documento")
        documento = str(valores[doc_idx]).strip() if doc_idx is not None and len(valores) > doc_idx else ""
        if documento.lower() in ("none", ""):
            documento = ""

        if cargo == 0 and abono == 0:
            continue

        resultado.filas.append(FilaCartola(
            fecha=fecha, descripcion=descripcion,
            cargo=cargo, abono=abono, saldo=saldo,
            documento=documento, fila_origen=i,
        ))

    resultado.total_filas_validas = len(resultado.filas)
    return resultado


# ─────────────────────────────────────────────
# Parser CSV genérico
# ─────────────────────────────────────────────

def parsear_csv(archivo_bytes: bytes, nombre_archivo: str = "") -> ResultadoImportacion:
    """Detecta el banco y enruta al parser correcto."""
    try:
        texto = decodificar(archivo_bytes)
    except ValueError:
        r = ResultadoImportacion()
        r.errores.append("No se pudo leer el archivo CSV.")
        return r

    banco = detectar_banco(texto[:600], nombre_archivo)

    if banco == "Global66":
        return parsear_global66(archivo_bytes, nombre_archivo)

    # CSV genérico
    resultado = ResultadoImportacion()
    resultado.banco_detectado = banco

    try:
        dialect = csv.Sniffer().sniff(texto[:2000], delimiters=",;\t|")
        reader = csv.reader(io.StringIO(texto), dialect)
    except Exception:
        reader = csv.reader(io.StringIO(texto))

    todas_filas = list(reader)
    resultado.total_filas_archivo = len(todas_filas)

    fila_enc_idx = None
    mapa_columnas = {}
    for i, fila in enumerate(todas_filas):
        if es_fila_encabezado(fila):
            mapa = detectar_columnas(fila)
            if "fecha" in mapa:
                fila_enc_idx = i
                mapa_columnas = mapa
                break

    if not mapa_columnas:
        resultado.errores.append("No se detectaron columnas en el CSV.")
        return resultado

    for i, fila in enumerate(todas_filas[fila_enc_idx + 1:], fila_enc_idx + 2):
        if not any(fila):
            continue
        fecha = parse_fecha(fila[mapa_columnas["fecha"]] if len(fila) > mapa_columnas.get("fecha", 99) else None)
        if fecha is None:
            continue
        desc_idx = mapa_columnas.get("descripcion")
        descripcion = fila[desc_idx].strip() if desc_idx is not None and len(fila) > desc_idx else "Movimiento bancario"
        cargo_idx = mapa_columnas.get("cargo")
        abono_idx = mapa_columnas.get("abono")
        cargo = parse_monto(fila[cargo_idx]) if cargo_idx is not None and len(fila) > cargo_idx else Decimal("0")
        abono = parse_monto(fila[abono_idx]) if abono_idx is not None and len(fila) > abono_idx else Decimal("0")
        if cargo == 0 and abono == 0:
            continue
        resultado.filas.append(FilaCartola(
            fecha=fecha, descripcion=descripcion,
            cargo=cargo, abono=abono, saldo=None, fila_origen=i,
        ))

    resultado.total_filas_validas = len(resultado.filas)
    return resultado


def _detectar_columnas_heuristico(todas_filas) -> Tuple[dict, int]:
    for i, fila in enumerate(todas_filas):
        for j, celda in enumerate(fila):
            if parse_fecha(celda) is not None:
                return {"fecha": j, "descripcion": j+2, "cargo": j+3, "abono": j+4, "saldo": j+5}, i
    return {}, 0


# ─────────────────────────────────────────────
# Clasificador genérico
# ─────────────────────────────────────────────

REGLAS_CLASIFICACION = [
    (["abono transferencia", "transferencia recibida", "deposito", "depósito",
      "venta", "factura", "boleta", "cobranza", "pago de cliente",
      "conversion de divisas", "conversión"], "ingreso_banco"),
    (["remuneracion", "remuneración", "sueldo", "salario",
      "liquidacion nomina", "pago personal", "planilla"], "remuneracion"),
    (["sii", "impuesto", "iva mensual", "renta", "patente",
      "contribucion", "contribución", "tesoreria", "tesorería"], "impuesto"),
    (["cuota prestamo", "cuota préstamo", "credito hipotecario",
      "leasing", "dividendo hipotecario"], "pago_prestamo"),
    (["comision", "comisión", "mantencion", "mantención",
      "cargo banco", "costo cuenta", "seguro"], "egreso_banco"),
    (["cheque", "pago cheque"], "egreso_banco"),
]


def clasificar_movimiento(descripcion: str, es_cargo: bool, tipo_banco: str = "") -> Optional[str]:
    tb = (tipo_banco or "").lower().strip()
    if tb in ("conversion_divisa", "transferencia_interna"):
        return tb
    if tipo_banco:
        resultado = clasificar_global66(tipo_banco, descripcion, es_cargo)
        if resultado is not None:
            return resultado

    desc = descripcion.lower()
    for palabras, tipo in REGLAS_CLASIFICACION:
        if any(p in desc for p in palabras):
            return tipo

    return "egreso_banco" if es_cargo else "ingreso_banco"


def convertir_a_movimientos(resultado: ResultadoImportacion) -> list:
    """
    Convierte las filas parseadas en dicts listos para guardar en MovimientoDiario.
    - monto: SIEMPRE en CLP
    - monto_moneda_orig: valor en moneda de la cuenta
    - tipo_cambio: TC usado al importar o estimado
    """
    movimientos = []
    for fila in resultado.filas:
        es_cargo = fila.cargo > 0
        tipo = clasificar_movimiento(fila.descripcion, es_cargo, fila.tipo_banco)
        if tipo is None:
            continue

        monto_orig = fila.cargo if es_cargo else fila.abono
        moneda = getattr(fila, "moneda", "CLP") or "CLP"
        tipo_cambio = getattr(fila, "tipo_cambio", None)
        if moneda != "CLP":
            tc = tipo_cambio if tipo_cambio and tipo_cambio > 0 else Decimal("950")
            monto_clp = monto_orig * tc
        else:
            tc = Decimal("1")
            monto_clp = monto_orig

        categoria_normalizada = "sin_clasificar"
        es_transferencia_interna = False
        if tipo == "conversion_divisa":
            categoria_normalizada = "conversion_divisa"
            es_transferencia_interna = True
            tipo = "egreso_banco" if es_cargo else "ingreso_banco"
        elif tipo == "transferencia_interna":
            categoria_normalizada = "transferencia_interna"
            es_transferencia_interna = True
            tipo = "egreso_banco" if es_cargo else "ingreso_banco"

        nota_doc = f" | Doc: {fila.documento}" if fila.documento else ""
        nota_tipo = f" | Tipo: {fila.tipo_banco}" if fila.tipo_banco else ""
        nota_moneda = f" | Moneda: {moneda}"
        nota_tc = f" | TC: {tc}" if tc else ""

        movimientos.append({
            "fecha": fila.fecha,
            "tipo": tipo,
            "descripcion": fila.descripcion[:250],
            "monto": monto_clp,
            "monto_moneda_orig": monto_orig,
            "moneda": moneda,
            "tipo_cambio": tc,
            "es_cargo": es_cargo,
            "medio_pago": "transferencia",
            "categoria_normalizada": categoria_normalizada,
            "es_transferencia_interna": es_transferencia_interna,
            "notas": f"Importado desde {resultado.banco_detectado}{nota_tipo}{nota_doc}{nota_moneda}{nota_tc}",
            "tipo_banco": fila.tipo_banco,
            "referencia_externa": fila.documento[:100] if fila.documento else "",
            "tercero": getattr(fila, "tercero", "")[:200],
            "pais_tercero": getattr(fila, "pais", "")[:5],
        })
    return movimientos
